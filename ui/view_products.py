import customtkinter as ctk
from tkinter import messagebox, filedialog
import csv
from fpdf import FPDF
from openpyxl import Workbook
from models.product import Product
from ui.edit_product import EditProductWindow

class ViewProductsWindow:
    def __init__(self, master, theme="System",  role="admin"):
        self.role = role
        self.window = ctk.CTkToplevel(master)
        self.window.title("All Products")
        self.window.geometry("800x600")
        self.window.resizable(False, False)
        
        ctk.set_appearance_mode(theme)

        # Title
        ctk.CTkLabel(self.window, text="All Products", font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(10, 5))

        # Search & Filter Frame
        filter_frame = ctk.CTkFrame(self.window)
        filter_frame.pack(fill="x", padx=15, pady=(0, 10))

        ctk.CTkLabel(filter_frame, text="Search:").pack(side="left", padx=(5, 5))
        self.search_var = ctk.StringVar()
        self.search_var.trace_add("write", lambda *args: self.load_products())
        self.search_entry = ctk.CTkEntry(filter_frame, textvariable=self.search_var, placeholder_text="Type product name...")
        self.search_entry.pack(side="left", padx=(0, 15))
        self.search_entry.bind("<KeyRelease>", lambda e: self.on_search_keyrelease())
        self.search_entry.focus_set()

        ctk.CTkLabel(filter_frame, text="Category:").pack(side="left", padx=(5, 5))
        categories = self.get_categories_from_db()
        self.category_var = ctk.StringVar(value="All")
        category_menu = ctk.CTkOptionMenu(filter_frame, variable=self.category_var, values=categories, 
                                         command=lambda _: self.load_products())
        category_menu.pack(side="left", padx=(0, 5))

        ctk.CTkButton(filter_frame, text="Reset Filters", command=self.reset_filters).pack(side="left", padx=5, pady=5)
        
        # 📤 Export Buttons
        export_frame = ctk.CTkFrame(self.window)
        export_frame.pack(fill="x", padx=15, pady=(0, 10))

        ctk.CTkButton(export_frame, text="Export CSV", command=self.export_csv).pack(side="left", padx=5)
        ctk.CTkButton(export_frame, text="Export Excel", command=self.export_excel).pack(side="left", padx=5)
        ctk.CTkButton(export_frame, text="Export PDF", command=self.export_pdf).pack(side="left", padx=5)

        # Table header
        header_frame = ctk.CTkFrame(self.window)
        header_frame.pack(fill="x", padx=15)

        headers = ["ID", "Name", "Category", "Quantity", "Price (₦)", "Action"]
        widths = [50, 180, 140, 100, 100, 100]
        for i, (title, width) in enumerate(zip(headers, widths)):
            lbl = ctk.CTkLabel(header_frame, text=title, font=ctk.CTkFont(size=14, weight="bold"))
            lbl.grid(row=0, column=i, padx=(0 if i == 0 else 10, 0), sticky="w")
            header_frame.grid_columnconfigure(i, minsize=width)

        # Scrollable frame for rows
        self.scroll_frame = ctk.CTkScrollableFrame(self.window, height=400)
        self.scroll_frame.pack(fill="both", expand=True, padx=15, pady=(5, 15))

        # Debug: Print initial products
        products = Product.get_all()
        print("Initial products:", [(p.id, p.name, p.category) for p in products])

        self.load_products()

    def get_categories_from_db(self):
        categories = Product.get_all_categories()  # Returns "No Category" for None/empty
        return categories  # Already normalized by Product.get_all_categories()

    def on_search_keyrelease(self):
        print("Search input (KeyRelease):", self.search_var.get(), "| Entry text:", self.search_entry.get())
        self.load_products()

    def load_products(self):
        for widget in self.scroll_frame.winfo_children():
            widget.destroy()

        # Fallback to direct entry.get() if StringVar fails
        search_text = self.search_var.get().strip().lower()
        if not search_text:  # Double-check with entry.get()
            search_text = self.search_entry.get().strip().lower()
        category_filter = self.category_var.get().strip()
        print(f"Applying filters: search='{search_text}', category='{category_filter}'")

        products = Product.get_all()
        print(f"Total products: {len(products)}")

        filtered_products = []
        for p in products:
            name = p.name.strip().lower() if isinstance(p.name, str) and p.name.strip() else ""
            category = p.category if isinstance(p.category, str) else "No Category"
            print(f"Processing product ID {p.id}: name='{p.name}', category='{category}'")

            if search_text and search_text not in name:
                print(f"Skipping product ID {p.id}: name '{p.name}' does not match search '{search_text}'")
                continue
            if category_filter != "All" and category != category_filter:
                print(f"Skipping product ID {p.id}: category '{category}' does not match filter '{category_filter}'")
                continue

            filtered_products.append(p)
            print(f"Including product ID {p.id}")

        print(f"Filtered products: {len(filtered_products)}")
        if not filtered_products:
            ctk.CTkLabel(self.scroll_frame, text="No matching products found", text_color="red").pack(pady=10)
            return

        LOW_STOCK_THRESHOLD = 5
        for p in filtered_products:
            category = p.category if isinstance(p.category, str) else "No Category"
            row_color = "#EB0000" if p.quantity <= LOW_STOCK_THRESHOLD else "transparent"
            row = ctk.CTkFrame(self.scroll_frame, fg_color=row_color)
            row.pack(fill="x", padx=5, pady=2)

            fields = [str(p.id), p.name, category, str(p.quantity), f"{p.price:,.2f}"]
            widths = [50, 180, 140, 100, 100]
            for i, (text, width) in enumerate(zip(fields, widths)):
                text_color = "#FFA500" if (i == 3 and p.quantity <= LOW_STOCK_THRESHOLD) else None
                label = ctk.CTkLabel(row, text=text, anchor="w", text_color=text_color)
                label.grid(row=0, column=i, sticky="w")
                row.grid_columnconfigure(i, minsize=width)

            edit_btn = ctk.CTkButton(
                row, text="✏ Edit", width=60, command=lambda pid=p.id: self.open_edit_window(pid),
                fg_color="#FFA500", hover_color="#FF8C00"
            )
            
            # 🔒 Disable for staff
            if self.role == "staff":
                edit_btn.configure(state="disabled")
                
            edit_btn.grid(row=0, column=len(fields), padx=5)

        self.window.update()

    def reset_filters(self):
        self.search_var.set("")
        self.search_entry.delete(0, "end")  # Explicitly clear entry
        self.category_var.set("All")
        self.load_products()

    def open_edit_window(self, product_id):
        EditProductWindow(self.window, product_id, refresh_callback=self.load_products)
        
    # ---------------- EXPORT FUNCTIONS ----------------
    def get_filtered_products(self):
        """Reapply filters and return filtered Product objects."""
        search_text = self.search_var.get().strip().lower()
        if not search_text:
            search_text = self.search_entry.get().strip().lower()
        category_filter = self.category_var.get().strip()

        products = Product.get_all()
        filtered_products = []
        for p in products:
            name = p.name.strip().lower() if isinstance(p.name, str) else ""
            category = p.category if isinstance(p.category, str) else "No Category"
            if search_text and search_text not in name:
                continue
            if category_filter != "All" and category != category_filter:
                continue
            filtered_products.append(p)
        return filtered_products

    def export_csv(self):
        products = self.get_filtered_products()
        if not products: return
        file = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")])
        if not file: return
        with open(file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Name", "Category", "Quantity", "Price"])
            for p in products:
                writer.writerow([p.id, p.name, p.category, p.quantity, p.price])
            messagebox.showinfo("Export", "CSV export successful!")

    def export_excel(self):
        products = self.get_filtered_products()
        if not products: return
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file: return
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Category", "Quantity", "Price"])
        for p in products:
            ws.append([p.id, p.name, p.category, p.quantity, p.price])
        wb.save(file)
        messagebox.showinfo("Export", "Excel export successful!")

    def export_pdf(self):
        products = self.get_filtered_products()
        if not products: return
        file = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
        if not file: return
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 10, "Products List", ln=True, align="C")
        pdf.ln(10)
        headers = ["ID", "Name", "Category", "Quantity", "Price"]
        for h in headers:
            pdf.cell(35, 8, h, border=1)
        pdf.ln()
        for p in products:
            pdf.cell(35, 8, str(p.id), border=1)
            pdf.cell(35, 8, str(p.name), border=1)
            pdf.cell(35, 8, str(p.category), border=1)
            pdf.cell(35, 8, str(p.quantity), border=1)
            pdf.cell(35, 8, f"{p.price:,.2f}", border=1)
            pdf.ln()
        pdf.output(file)
        messagebox.showinfo("Export successful!", f"Products exported to {file}")

        
        
    

